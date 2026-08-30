from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook


MONTHS = {
    "JAN": 1, "JANEIRO": 1, "FEV": 2, "FEVEREIRO": 2,
    "MAR": 3, "MARCO": 3, "ABR": 4, "ABRIL": 4,
    "MAI": 5, "MAIO": 5, "JUN": 6, "JUNHO": 6,
    "JUL": 7, "JULHO": 7, "AGO": 8, "AGOSTO": 8,
    "SET": 9, "SETEMBRO": 9, "OUT": 10, "OUTUBRO": 10,
    "NOV": 11, "NOVEMBRO": 11, "DEZ": 12, "DEZEMBRO": 12,
}


def normalized(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return " ".join(
        "".join(char for char in text if not unicodedata.combining(char))
        .upper().strip().split()
    )


def decimal_value(value: object) -> Decimal:
    if isinstance(value, float):
        value = str(value)
    try:
        result = Decimal(value).quantize(Decimal("0.0001"))
    except (InvalidOperation, TypeError, ValueError) as error:
        raise ValueError("Valor numérico inválido.") from error
    if not result.is_finite() or result < 0:
        raise ValueError("Valor numérico inválido.")
    return result


@dataclass(slots=True)
class HistoricalPreview:
    file_path: Path
    detected_type: str
    year: int | None
    rows: list[dict] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    duplicates: int = 0
    total: Decimal = Decimal("0.0000")

    @property
    def valid_rows(self) -> int:
        return sum(
            not row.get("duplicate") and not row.get("skip") for row in self.rows
        )

    @property
    def can_import(self) -> bool:
        return bool(self.rows) and not self.errors


class HistoricalWorkbookImporter:
    """Leitor conservador. Nunca persiste dados e nunca altera a planilha."""

    def parse(self, file_path: str | Path) -> HistoricalPreview:
        path = Path(file_path)
        if path.suffix.lower() not in {".xlsx", ".xlsm"} or not path.is_file():
            return HistoricalPreview(path, "DESCONHECIDO", None,
                                     errors=["Arquivo Excel inválido ou inexistente."])
        workbook = load_workbook(path, read_only=True, data_only=True)
        names = {normalized(name): name for name in workbook.sheetnames}
        if "LANCAMENTOS" in names:
            return self._cashflow(path, workbook[names["LANCAMENTOS"]])
        if "META" in names and "FATURAMENTO" in names:
            preview = self._targets(path, workbook[names["META"]], workbook[names["FATURAMENTO"]])
            if "ASSOCIACOES" in names:
                preview.warnings.append(
                    "A planilha também contém Associação; selecione esse tipo no preview para importá-la."
                )
            return preview
        if "PLANEJ. ORCAMENTARIO" in names:
            return self._budget(path, workbook[names["PLANEJ. ORCAMENTARIO"]])
        if "TAXA BOE" in names:
            return HistoricalPreview(path, "BOE", None)
        return HistoricalPreview(path, "DESCONHECIDO", None,
                                 errors=["Estrutura oficial não reconhecida."])

    def parse_association(self, file_path: str | Path) -> HistoricalPreview:
        path = Path(file_path)
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_name = next((name for name in workbook.sheetnames
                           if normalized(name) == "ASSOCIACOES"), None)
        if sheet_name is None:
            return HistoricalPreview(path, "ASSOCIACAO", None,
                                     errors=["A aba Associações não foi encontrada."])
        sheet = workbook[sheet_name]
        preview = HistoricalPreview(path, "ASSOCIACAO", self._year_from_name(path) or 2026)
        rows = sheet.iter_rows(values_only=True)
        next(rows, None)
        next(rows, None)
        for line, values in enumerate(rows, start=3):
            code = values[0] if values else None
            if not isinstance(code, (int, float)) or int(code) < 7501:
                continue
            if int(code) == 7500:
                continue
            for month in range(1, 13):
                base = 2 + (month - 1) * 4
                cancellation = values[base] if len(values) > base else None
                capture = values[base + 1] if len(values) > base + 1 else None
                execution = values[base + 3] if len(values) > base + 3 else None
                if cancellation is None and capture is None and execution is None:
                    continue
                try:
                    preview.rows.append({"line": line, "code": int(code),
                                         "year": preview.year, "month": month,
                                         "cancellation": decimal_value(cancellation or 0),
                                         "capture": decimal_value(capture or 0),
                                         "execution": decimal_value(execution or 0)})
                except ValueError as error:
                    preview.errors.append(f"Linha {line}, mês {month}: {error}")
        if not preview.rows:
            preview.errors.append("Nenhum dado mensal de Associação foi encontrado.")
        return preview

    def _cashflow(self, path: Path, sheet) -> HistoricalPreview:
        preview = HistoricalPreview(path, "FLUXO_CAIXA", None)
        for line, values in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            if not any(value is not None for value in values[:8]):
                continue
            year, month_name, description, notes, category, kind, value, boe = values[:8]
            if description is None and "SALDO" in normalized(notes):
                preview.warnings.append(f"Linha {line}: saldo técnico não importado como lançamento.")
                continue
            try:
                year = int(year)
                month = MONTHS[normalized(month_name)]
                amount = decimal_value(value)
                if amount <= 0:
                    raise ValueError("O valor deve ser maior que zero.")
                if not str(description or "").strip():
                    raise ValueError("Descrição obrigatória.")
                if normalized(boe) not in {"SIM", "NAO"}:
                    raise ValueError("BOE deve ser Sim ou Não.")
                row = {"line": line, "year": year, "month": month,
                       "description": str(description).strip(),
                       "notes": str(notes).strip() if notes is not None else None,
                       "category_label": str(category or "").strip(),
                       "type_label": str(kind or "").strip(), "value": amount,
                       "boe": normalized(boe) == "SIM"}
                preview.rows.append(row)
                preview.total += amount
                preview.year = year if preview.year is None else preview.year
            except (KeyError, TypeError, ValueError) as error:
                preview.errors.append(f"Linha {line}: {error}")
        return preview

    def _targets(self, path: Path, target_sheet, actual_sheet) -> HistoricalPreview:
        preview = HistoricalPreview(path, "META_REALIZADO", self._year_from_name(path) or 2026)
        actual_by_key = {}
        for values in actual_sheet.iter_rows(min_row=4, values_only=True):
            if isinstance(values[0], (int, float)) and int(values[0]) >= 7501:
                for month in range(1, 13):
                    actual_by_key[(int(values[0]), month)] = (
                        values[month + 1] if len(values) > month + 1 else None
                    )
        for line, values in enumerate(target_sheet.iter_rows(min_row=4, values_only=True), start=4):
            if not isinstance(values[0], (int, float)) or int(values[0]) < 7501:
                continue
            code = int(values[0])
            if code == 7500:
                continue
            for month in range(1, 13):
                target = values[month + 1] if len(values) > month + 1 else None
                actual = actual_by_key.get((code, month))
                if target is None and actual is None:
                    continue
                try:
                    preview.rows.append({"line": line, "code": code,
                                         "year": preview.year, "month": month,
                                         "indicator": "CONSULTAS",
                                         "target": decimal_value(target or 0),
                                         "actual": decimal_value(actual or 0)})
                except ValueError as error:
                    preview.errors.append(f"Linha {line}, mês {month}: {error}")
        preview.warnings.append(
            "A estrutura analisada contém CONSULTAS. REGISTROS só serão importados quando houver aba oficial inequívoca."
        )
        return preview

    def _budget(self, path: Path, sheet) -> HistoricalPreview:
        year = self._year_from_name(path) or 2026
        preview = HistoricalPreview(path, "ORCAMENTO", year)
        mappings = {
            "REPASSE CDLS ESTADO GOIAS": ("RECEITA", "RECEITA_DIRETA"),
            "DESPESAS COM PESSOAL": ("DESPESA", "PESSOAL"),
            "DESPESAS COM EVENTOS": ("DESPESA", "EVENTOS"),
            "DESPESAS COM A OPERACAO": ("DESPESA", "OPERACIONAL"),
        }
        ignored = 0
        for line, values in enumerate(sheet.iter_rows(min_row=9, values_only=True), start=9):
            label = normalized(values[1] if len(values) > 1 else None)
            mapping = mappings.get(label)
            if mapping is None:
                if label and any(values[index] for index in range(2, min(len(values), 26), 2)):
                    ignored += 1
                continue
            for month in range(1, 13):
                value = values[2 + (month - 1) * 2]
                if value is None:
                    continue
                amount = decimal_value(value)
                preview.rows.append({"line": line, "year": year, "month": month,
                                     "type": mapping[0], "category": mapping[1],
                                     "value": amount, "source_label": values[1]})
                preview.total += amount
        if ignored:
            preview.warnings.append(
                f"{ignored} linhas detalhadas sem correspondência inequívoca foram excluídas do preview."
            )
        return preview

    @staticmethod
    def _year_from_name(path: Path) -> int | None:
        match = re.search(r"\b(20\d{2})\b", path.name)
        return int(match.group(1)) if match else None
