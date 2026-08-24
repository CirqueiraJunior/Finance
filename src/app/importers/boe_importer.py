from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import hashlib
from pathlib import Path
import re
import unicodedata
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.importers.boe_types import (
    BOEIssueSeverity,
    BOEParsedRow,
    BOEValidationIssue,
    BOEValidationResult,
)


class BOEImporter:
    SHEET_NAME = "Taxa BOE"
    ALLOWED_EXTENSIONS = {".xlsx"}
    HEADER_LABELS = {"ROW LABELS", "ENTIDADE"}
    HEADER_QUANTITY = "QTDE DE CONSULTAS"
    HEADER_VALUE = "VALOR TOTAL"
    CONSOLIDATED_CODE = 7500
    MONEY_QUANTUM = Decimal("0.0001")

    def parse(self, file_path: str | Path) -> BOEValidationResult:
        path = Path(file_path).expanduser().resolve()
        result = BOEValidationResult(caminho_arquivo=path, nome_arquivo=path.name)

        if not path.is_file():
            self._add_error(result, "Arquivo BOE não encontrado.")
            return result
        if path.suffix.lower() not in self.ALLOWED_EXTENSIONS:
            self._add_error(result, "Extensão inválida. Utilize um arquivo .xlsx.")
            return result

        result.hash_arquivo = self._calculate_hash(path)
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except (InvalidFileException, OSError, ValueError) as error:
            self._add_error(result, f"Não foi possível abrir o workbook: {error}")
            return result

        try:
            if self.SHEET_NAME not in workbook.sheetnames:
                self._add_error(result, 'A aba obrigatória "Taxa BOE" não foi encontrada.')
                return result
            sheet = workbook[self.SHEET_NAME]
            period = self._extract_period_from_sheet(sheet)
            if period is None:
                period = self._extract_period_from_filename(path.name)
            if period is None:
                self._add_error(result, "Não foi possível identificar o período do BOE.")
            else:
                result.periodo_mes, result.periodo_ano = period

            header = self._find_header(sheet)
            if header is None:
                self._add_error(result, "Estrutura inválida na aba Taxa BOE: cabeçalhos ausentes.")
                return result
            pivot_mapping: dict[str, int] = {}
            if header[4] == "condensed":
                pivot_mapping = self._read_pivot_entity_mapping(path)
                if not pivot_mapping:
                    self._add_error(
                        result,
                        "A tabela dinâmica não disponibiliza o mapeamento entre Entidade e código.",
                    )
                    return result
            self._parse_rows(sheet, header, result, pivot_mapping)
            if not result.linhas:
                self._add_error(result, "Nenhuma linha de Entidade válida foi encontrada.")
            self._validate_consolidated(result)
            return result
        finally:
            workbook.close()

    def _find_header(self, sheet: object) -> tuple[int, int, int, int, str] | None:
        for row_number, row in enumerate(
            sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20), values_only=True),
            start=1,
        ):
            normalized = [self._normalize_text(value) for value in row]
            try:
                label_column = next(
                    index + 1
                    for index, value in enumerate(normalized)
                    if value in self.HEADER_LABELS
                )
                quantity_column = normalized.index(self.HEADER_QUANTITY) + 1
                value_column = normalized.index(self.HEADER_VALUE) + 1
            except (StopIteration, ValueError):
                continue
            mode = "hierarchical" if normalized[label_column - 1] == "ROW LABELS" else "condensed"
            return row_number, label_column, quantity_column, value_column, mode
        return None

    def _parse_rows(
        self,
        sheet: object,
        header: tuple[int, int, int, int, str],
        result: BOEValidationResult,
        pivot_mapping: dict[str, int],
    ) -> None:
        header_row, label_column, quantity_column, value_column, mode = header
        pending_name: tuple[int, str] | None = None
        consolidated_reported = False

        for row_number, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            label = self._cell(row, label_column)
            quantity = self._cell(row, quantity_column)
            value = self._cell(row, value_column)
            label_text = str(label).strip() if label is not None else ""
            normalized_label = self._normalize_text(label)

            if not label_text and quantity is None and value is None:
                continue
            if normalized_label == "GRAND TOTAL":
                pending_name = None
                continue
            if normalized_label == "TOTAL GERAL":
                self._capture_consolidated(
                    result, row_number, quantity, value, codigo=None
                )
                consolidated_reported = True
                pending_name = None
                continue
            if normalized_label in {"(VAZIO)", "VAZIO"}:
                pending_name = None
                continue
            if not label_text and (quantity is not None or value is not None):
                if not consolidated_reported:
                    self._capture_consolidated(
                        result, row_number, quantity, value, codigo=None
                    )
                    consolidated_reported = True
                pending_name = None
                continue

            if mode == "condensed":
                code = pivot_mapping.get(normalized_label)
                if code is None:
                    self._add_error(
                        result,
                        "Não foi possível identificar o código da Entidade na tabela dinâmica.",
                        linha=row_number,
                    )
                    continue
                if code == self.CONSOLIDATED_CODE:
                    self._capture_consolidated(
                        result, row_number, quantity, value, codigo=str(code)
                    )
                    consolidated_reported = True
                    continue
                parsed_quantity = self._parse_quantity(
                    quantity, result, row_number, str(code)
                )
                parsed_value = self._parse_money(
                    value, result, row_number, str(code)
                )
                if parsed_quantity is not None and parsed_value is not None:
                    result.linhas.append(
                        BOEParsedRow(
                            linha=row_number,
                            codigo_entidade=code,
                            nome_entidade=label_text,
                            quantidade_consultas=parsed_quantity,
                            valor_total=parsed_value,
                        )
                    )
                continue

            code = self._parse_code(label)
            if code is None:
                if pending_name is not None:
                    self._add_error(
                        result,
                        "Nome de Entidade sem linha de código correspondente.",
                        linha=pending_name[0],
                    )
                pending_name = (row_number, label_text)
                continue

            if code == self.CONSOLIDATED_CODE:
                self._capture_consolidated(
                    result, row_number, quantity, value, codigo=str(code)
                )
                consolidated_reported = True
                pending_name = None
                continue
            if pending_name is None:
                self._add_error(
                    result,
                    "Código de Entidade sem nome correspondente na linha anterior.",
                    linha=row_number,
                    codigo=str(code),
                )
                continue

            parsed_quantity = self._parse_quantity(
                quantity, result, row_number, str(code)
            )
            parsed_value = self._parse_money(value, result, row_number, str(code))
            if parsed_quantity is not None and parsed_value is not None:
                result.linhas.append(
                    BOEParsedRow(
                        linha=row_number,
                        codigo_entidade=code,
                        nome_entidade=pending_name[1],
                        quantidade_consultas=parsed_quantity,
                        valor_total=parsed_value,
                    )
                )
            pending_name = None

        if pending_name is not None:
            self._add_error(
                result,
                "Nome de Entidade sem linha de código correspondente.",
                linha=pending_name[0],
            )

    @classmethod
    def _read_pivot_entity_mapping(cls, path: Path) -> dict[str, int]:
        namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        try:
            with ZipFile(path) as archive:
                definition_names = sorted(
                    name
                    for name in archive.namelist()
                    if name.startswith("xl/pivotCache/pivotCacheDefinition")
                    and name.endswith(".xml")
                )
                for definition_name in definition_names:
                    root = ElementTree.fromstring(archive.read(definition_name))
                    fields = root.find("x:cacheFields", namespace)
                    if fields is None:
                        continue
                    by_name = {
                        field.get("name"): field
                        for field in fields.findall("x:cacheField", namespace)
                    }
                    code_field = by_name.get("COD_ENTIDADE_COBRANCA")
                    name_field = by_name.get("NOM_ENTIDADE_COBRANCA")
                    if code_field is None or name_field is None:
                        continue
                    codes = cls._read_shared_items(code_field, namespace)
                    names = cls._read_shared_items(name_field, namespace)
                    if len(codes) != len(names):
                        continue
                    mapping: dict[str, int] = {}
                    for code, entity_name in zip(codes, names, strict=True):
                        if code is None or entity_name is None:
                            continue
                        try:
                            parsed_code = int(Decimal(str(code)))
                        except (InvalidOperation, TypeError, ValueError):
                            continue
                        normalized_name = cls._normalize_text(entity_name)
                        if normalized_name:
                            mapping[normalized_name] = parsed_code
                    if mapping:
                        return mapping
        except (BadZipFile, KeyError, OSError, ElementTree.ParseError):
            return {}
        return {}

    @staticmethod
    def _read_shared_items(
        field: ElementTree.Element,
        namespace: dict[str, str],
    ) -> list[str | None]:
        shared_items = field.find("x:sharedItems", namespace)
        if shared_items is None:
            return []
        values: list[str | None] = []
        for item in list(shared_items):
            tag = item.tag.rsplit("}", 1)[-1]
            values.append(None if tag == "m" else item.get("v"))
        return values

    def _capture_consolidated(
        self,
        result: BOEValidationResult,
        row_number: int,
        quantity: object,
        value: object,
        codigo: str | None,
    ) -> None:
        parsed_quantity = self._parse_quantity(quantity, result, row_number, codigo)
        parsed_value = self._parse_money(value, result, row_number, codigo)
        if parsed_quantity is not None and parsed_value is not None:
            result.quantidade_consolidada = parsed_quantity
            result.valor_consolidado = parsed_value
        result.inconsistencias.append(
            BOEValidationIssue(
                mensagem="Linha consolidada ignorada; não representa uma Entidade real.",
                severidade=BOEIssueSeverity.WARNING,
                linha=row_number,
                codigo=codigo,
            )
        )

    def _validate_consolidated(self, result: BOEValidationResult) -> None:
        if result.quantidade_consolidada is None or result.valor_consolidado is None:
            return
        calculated_quantity = sum(row.quantidade_consultas for row in result.linhas)
        calculated_value = result.valor_total_calculado
        if calculated_quantity != result.quantidade_consolidada:
            self._add_error(
                result,
                "A quantidade consolidada não confere com a soma das Entidades.",
            )
        if abs(calculated_value - result.valor_consolidado) > self.MONEY_QUANTUM:
            self._add_error(
                result,
                "O valor consolidado não confere com a soma das Entidades.",
            )

    @classmethod
    def _parse_quantity(
        cls,
        value: object,
        result: BOEValidationResult,
        row_number: int,
        code: str | None,
    ) -> int | None:
        try:
            decimal_value = Decimal(str(value))
            if not decimal_value.is_finite() or decimal_value != decimal_value.to_integral_value():
                raise InvalidOperation
            quantity = int(decimal_value)
            if quantity < 0:
                raise InvalidOperation
            return quantity
        except (InvalidOperation, TypeError, ValueError):
            cls._add_error(
                result,
                "Quantidade de consultas inválida.",
                linha=row_number,
                codigo=code,
            )
            return None

    @classmethod
    def _parse_money(
        cls,
        value: object,
        result: BOEValidationResult,
        row_number: int,
        code: str | None,
    ) -> Decimal | None:
        try:
            amount = Decimal(str(value))
            if not amount.is_finite() or amount < 0:
                raise InvalidOperation
            return amount.quantize(cls.MONEY_QUANTUM)
        except (InvalidOperation, TypeError, ValueError):
            cls._add_error(
                result,
                "Valor total inválido.",
                linha=row_number,
                codigo=code,
            )
            return None

    @staticmethod
    def _parse_code(value: object) -> int | None:
        text = str(value).strip()
        if not re.fullmatch(r"\d+", text):
            return None
        return int(text)

    @classmethod
    def _extract_period_from_sheet(cls, sheet: object) -> tuple[int, int] | None:
        pattern = re.compile(
            r"(?:COMPETENCIA|PERIODO|REFERENCIA)[^0-9]*"
            r"(0?[1-9]|1[0-2])[/._-](20\d{2}|\d{2})"
        )
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10), values_only=True):
            for value in row:
                if isinstance(value, (datetime, date)):
                    return value.month, value.year
                normalized = cls._normalize_text(value)
                match = pattern.search(normalized)
                if match:
                    return cls._normalize_period(match.group(1), match.group(2))
        return None

    @classmethod
    def _extract_period_from_filename(cls, filename: str) -> tuple[int, int] | None:
        match = re.search(
            r"(?<!\d)(0[1-9]|1[0-2])[._-](\d{2}|\d{4})(?!\d)", filename
        )
        if not match:
            return None
        return cls._normalize_period(match.group(1), match.group(2))

    @staticmethod
    def _normalize_period(month: str, year: str) -> tuple[int, int]:
        normalized_year = int(year)
        if len(year) == 2:
            normalized_year += 2000
        return int(month), normalized_year

    @staticmethod
    def _calculate_hash(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as stream:
            for chunk in iter(lambda: stream.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _cell(row: tuple[object, ...], one_based_column: int) -> object:
        index = one_based_column - 1
        return row[index] if index < len(row) else None

    @staticmethod
    def _normalize_text(value: object) -> str:
        if value is None:
            return ""
        text = unicodedata.normalize("NFKD", str(value))
        without_accents = "".join(char for char in text if not unicodedata.combining(char))
        return " ".join(without_accents.upper().strip().split())

    @staticmethod
    def _add_error(
        result: BOEValidationResult,
        message: str,
        *,
        linha: int | None = None,
        codigo: str | None = None,
    ) -> None:
        result.inconsistencias.append(
            BOEValidationIssue(
                mensagem=message,
                severidade=BOEIssueSeverity.ERROR,
                linha=linha,
                codigo=codigo,
            )
        )
