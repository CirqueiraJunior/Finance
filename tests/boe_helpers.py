from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook

from app.models.boe_import import BOEImport
from sqlalchemy.orm import Session


DEFAULT_ROWS = [
    (7501, "CDL GOIANIA/GO", 100, Decimal("6.9300")),
    (7544, "CDL ANAPOLIS/GO", 50, Decimal("3.4650")),
]


def create_boe_workbook(
    path: Path,
    *,
    rows: list[tuple[int, str, int, Decimal]] | None = None,
    sheet_name: str = "Taxa BOE",
    valid_headers: bool = True,
    include_consolidated: bool = False,
    explicit_period: str | None = None,
    blank_rows: bool = False,
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    if explicit_period:
        sheet.cell(1, 1, explicit_period)
    sheet.cell(2, 1, "REPASSE BOE - GOIÁS")
    if valid_headers:
        sheet.append([])
        sheet.append(["Row Labels", "Qtde de Consultas", "Valor Total"])
    else:
        sheet.append([])
        sheet.append(["Coluna A", "Coluna B", "Coluna C"])

    selected_rows = DEFAULT_ROWS if rows is None else rows
    for code, name, quantity, amount in selected_rows:
        sheet.append([name, quantity, float(amount)])
        if blank_rows:
            sheet.append([None, None, None])
        sheet.append([str(code), quantity, float(amount)])
    if include_consolidated:
        quantity = sum(row[2] for row in selected_rows)
        amount = sum((row[3] for row in selected_rows), start=Decimal("0"))
        sheet.append(["7500", quantity, float(amount)])
    sheet.append(["Grand Total", None, None])
    workbook.save(path)
    workbook.close()
    return path


def add_boe_import(
    session: Session,
    *,
    year: int = 2026,
    month: int = 7,
    file_hash: str = "a" * 64,
    filename: str = "BOE - 07.26.xlsx",
) -> BOEImport:
    boe_import = BOEImport(
        periodo_ano=year,
        periodo_mes=month,
        nome_arquivo=filename,
        caminho_origem=f"C:/imports/{filename}",
        hash_arquivo=file_hash,
        quantidade_entidades=0,
        quantidade_inconsistencias=0,
        valor_total=Decimal("0.0000"),
        status="imported",
    )
    session.add(boe_import)
    session.commit()
    return boe_import

