from decimal import Decimal
from pathlib import Path
import sqlite3

from openpyxl import Workbook
import pytest
from PySide6.QtCore import Qt, qInstallMessageHandler
from PySide6.QtWidgets import QLabel, QPushButton, QStackedWidget
from sqlalchemy import create_engine, func, select
from sqlalchemy.orm import Session

from app.core.config import Settings
from app.core.version import __version__
from app.core.exceptions import EntityDomainError
from app.gui.pages.administracao import AdministracaoPage
from app.gui.pages.cadastros import CadastrosPage
from app.gui.pages.dashboard import DashboardPage
from app.gui.main_window import MainWindow
from app.core.config import get_settings
from app.resources import load_stylesheet
from app.importers.historical_importer import HistoricalPreview, HistoricalWorkbookImporter
from app.models.budget_entry import BudgetEntry
from app.models.cashflow_catalog_entry import CashflowCatalogEntry
from app.models.entity import Entity
from app.gui.controllers.navigation_controller import NavigationController
from app.widgets.sidebar import Sidebar
from app.repositories.cashflow_catalog_repository import CashflowCatalogRepository
from app.repositories.entity_repository import EntityRepository
from app.services.backup_service import BackupService
from app.services.cashflow_catalog_service import CashflowCatalogService
from app.services.entity_service import EntityService
from app.services.historical_import_service import HistoricalImportService


def test_release_has_single_version_source():
    assert __version__ == "1.0.0"


def test_entity_maintenance_updates_and_toggles_without_allowing_7500(db_session):
    service = EntityService(EntityRepository(db_session))
    with pytest.raises(EntityDomainError, match="7500"):
        service.create_entity(codigo_entidade=7500, nome="Consolidado")
    entity = service.create_entity(codigo_entidade=7501, nome="Goiânia")
    service.update_entity(
        entity.id, nome="Goiânia", nome_oficial="CDL Goiânia",
        municipio="Goiânia", uf="go", sigla="GYN", ativa=True,
    )
    assert entity.nome_oficial == "CDL Goiânia"
    assert entity.uf == "GO"
    service.set_active(entity.id, False)
    assert not entity.ativa


def test_catalog_maintenance_rejects_incoherent_combinations(db_session):
    service = CashflowCatalogService(CashflowCatalogRepository(db_session))
    entry = service.create_entry(
        description="Telefone", category="ADMINISTRATIVO",
        movement_type="DESPESA",
    )
    assert entry in service.list_entries()
    with pytest.raises(ValueError, match="incoerente"):
        service.create_entry(
            description="Inválido", category="PESSOAL", movement_type="RECEITA"
        )


def test_registration_page_exposes_entities_and_catalog_actions(qtbot):
    page = CadastrosPage()
    qtbot.addWidget(page)
    assert page.tabs.count() == 2
    assert page.new_entity_button.text() == "Nova Entidade"
    assert page.new_catalog_button.text() == "Novo Item"
    assert page.entity_table.columnCount() == 7


def test_administration_page_buttons_respond_to_mouse(qtbot):
    page = AdministracaoPage()
    qtbot.addWidget(page)
    page.show()
    clicks = []
    page.backup_button.clicked.connect(lambda: clicks.append("backup"))
    page.refresh_button.clicked.connect(lambda: clicks.append("refresh"))
    qtbot.mouseClick(page.backup_button, Qt.MouseButton.LeftButton)
    qtbot.mouseClick(page.refresh_button, Qt.MouseButton.LeftButton)
    assert clicks == ["backup", "refresh"]


def test_sqlite_manual_backup_is_valid_and_never_overwrites(tmp_path):
    database = tmp_path / "source.db"
    with sqlite3.connect(database) as connection:
        connection.execute("CREATE TABLE evidence (value TEXT)")
        connection.execute("INSERT INTO evidence VALUES ('ok')")
    settings = Settings("J.A. Finance", "test", False,
                        f"sqlite:///{database.as_posix()}", "INFO", tmp_path / "logs")
    service = BackupService(settings)
    service._create = lambda directory: BackupService._create(service, tmp_path / directory.name)
    first = service.create_manual_backup()
    second = service.create_manual_backup()
    assert first != second
    with sqlite3.connect(first) as connection:
        assert connection.execute("SELECT value FROM evidence").fetchone()[0] == "ok"


def test_cashflow_preview_skips_empty_rows_and_technical_balance(tmp_path):
    path = tmp_path / "fluxo.xlsm"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Lançamentos"
    sheet.append([])
    sheet.append(["Ano", "Mês", "Descrição", "Observação", "Categoria", "Tipo", "Valor", "BOE"])
    sheet.append([2026, "JUL", None, "Saldo Inicial", None, None, 100, "Não"])
    sheet.append([2026, "JUL", "Telefone", None, "Administrativo", "Despesa", 10.5, "Não"])
    sheet.append([])
    workbook.save(path)
    preview = HistoricalWorkbookImporter().parse(path)
    assert preview.detected_type == "FLUXO_CAIXA"
    assert len(preview.rows) == 1
    assert preview.rows[0]["value"] == Decimal("10.5000")
    assert preview.warnings


def test_budget_preview_only_maps_unambiguous_aggregate_rows(tmp_path):
    path = tmp_path / "Projeção Orçamentária 2026.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Planej. orçamentário"
    for _ in range(8):
        sheet.append([])
    row = [None, "Despesas com Pessoal"]
    for month in range(12):
        row.extend([100 + month, None])
    sheet.append(row)
    detailed = [None, "Item sem mapeamento"] + [50, None] * 12
    sheet.append(detailed)
    workbook.save(path)
    preview = HistoricalWorkbookImporter().parse(path)
    assert preview.detected_type == "ORCAMENTO"
    assert len(preview.rows) == 12
    assert {row["category"] for row in preview.rows} == {"PESSOAL"}
    assert preview.warnings


def test_target_and_association_previews_exclude_7500(tmp_path):
    path = tmp_path / "Meta x Realizado - Oficial 2026.xlsm"
    workbook = Workbook()
    target = workbook.active
    target.title = "Meta"
    target.append([])
    target.append([])
    target.append(["COD.", "Entidade", "JAN"])
    target.append([7500, "Consolidado", 999])
    target.append([7501, "Goiânia", 100])
    actual = workbook.create_sheet("Faturamento")
    actual.append([])
    actual.append([])
    actual.append(["COD.", "Entidade", "JAN"])
    actual.append([7500, "Consolidado", 999])
    actual.append([7501, "Goiânia", 80])
    association = workbook.create_sheet("Associações")
    association.append(["CÓD", "ENTIDADES", "JANEIRO", None, None, None])
    association.append([None, None, "CANC.", "CAPTAÇÃO", "SUSPENSO", "TOTAL ASSC."])
    association.append([7500, "Consolidado", 0, 99, 0, 999])
    association.append([7501, "Goiânia", 0, 10, 0, 50])
    workbook.save(path)

    importer = HistoricalWorkbookImporter()
    targets = importer.parse(path)
    associations = importer.parse_association(path)
    assert {(row["code"], row["target"], row["actual"]) for row in targets.rows} == {
        (7501, Decimal("100.0000"), Decimal("80.0000"))
    }
    assert {(row["code"], row["capture"], row["execution"])
            for row in associations.rows} == {
        (7501, Decimal("10.0000"), Decimal("50.0000"))
    }


def test_budget_deduplication_marks_existing_functional_key(db_session, tmp_path):
    db_session.add(BudgetEntry(
        periodo_ano=2026, periodo_mes=1, tipo="DESPESA", categoria="PESSOAL",
        valor_orcado=Decimal("100"),
    ))
    db_session.commit()
    service = HistoricalImportService(
        db_session, HistoricalWorkbookImporter(), None, None, None, None
    )
    preview = HistoricalPreview(tmp_path / "budget.xlsx", "ORCAMENTO", 2026, rows=[
        {"year": 2026, "month": 1, "type": "DESPESA", "category": "PESSOAL",
         "value": Decimal("100"), "source_label": "Pessoal"},
    ])
    service._mark_duplicates(preview)
    assert preview.duplicates == 1
    assert preview.rows[0]["duplicate"] is True


def test_import_rolls_back_everything_when_one_row_fails(db_session, tmp_path, monkeypatch):
    class BackupStub:
        def create_import_backup(self):
            path = tmp_path / "backup.db"
            path.write_bytes(b"backup")
            return path

    service = HistoricalImportService(
        db_session, HistoricalWorkbookImporter(), None, None, BackupStub(), None
    )
    preview = HistoricalPreview(tmp_path / "budget.xlsx", "ORCAMENTO", 2026, rows=[
        {"year": 2026, "month": 1, "type": "DESPESA", "category": "PESSOAL",
         "value": Decimal("100"), "source_label": "Pessoal"},
        {"year": 2026, "month": 2, "type": "DESPESA", "category": "PESSOAL",
         "value": Decimal("200"), "source_label": "Pessoal"},
    ])
    original = service._persist
    calls = 0

    def failing(kind, row):
        nonlocal calls
        calls += 1
        if calls == 2:
            raise RuntimeError("falha simulada")
        original(kind, row)

    monkeypatch.setattr(service, "_persist", failing)
    with pytest.raises(RuntimeError, match="simulada"):
        service.import_preview(preview)
    assert db_session.scalar(select(func.count()).select_from(BudgetEntry)) == 0


def test_qfont_initialization_emits_no_invalid_point_size(qtbot, qapp):
    messages = []

    def handler(_type, _context, message):
        messages.append(message)

    previous = qInstallMessageHandler(handler)
    previous_stylesheet = qapp.styleSheet()
    try:
        qapp.setStyleSheet(load_stylesheet())
        window = MainWindow(get_settings())
        qtbot.addWidget(window)
        window.show()
        qapp.processEvents()
    finally:
        qapp.setStyleSheet(previous_stylesheet)
        qInstallMessageHandler(previous)
    assert not any("Point size <= 0" in message for message in messages)


def test_all_sidebar_buttons_navigate_with_real_mouse_click(qtbot):
    stack = QStackedWidget()
    qtbot.addWidget(stack)
    indexes = {key: stack.addWidget(QLabel(label)) for label, key in Sidebar.ITEMS}
    controller = NavigationController(stack, indexes)
    sidebar = Sidebar(controller.navigate_to)
    qtbot.addWidget(sidebar)
    sidebar.show()
    for label, key in Sidebar.ITEMS:
        button = next(
            item for item in sidebar.findChildren(QPushButton)
            if item.property("pageKey") == key
        )
        qtbot.mouseClick(button, Qt.MouseButton.LeftButton)
        assert stack.currentIndex() == indexes[key], label
